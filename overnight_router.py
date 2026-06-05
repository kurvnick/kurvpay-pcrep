#!/usr/bin/env python3
"""
overnight_router.py  —  Overnight lead-routing engine (v2), full-population decision log.

WHAT IT DOES
  Pulls every Lead created in a date window from Zoho CRM, keeps the ones created during the
  overnight window (default 6:00pm -> 8:30am PT), runs the exact v2 routing logic on each night's
  batch (parity is per-night), and writes an Excel with two tabs:
     - "Lead Decisions": one row per overnight lead — decision, assignee, why, flags, eventual status
     - "Night Summary":  decision mix per night
  Run it once with a 90-day window for the backfill, then daily (cron) for the prior night.

WHY THIS EXISTS
  ~9,000 overnight leads over 90 days can't be round-tripped through a chat reliably. This runs
  where the data lives, so the log is complete and the contact fields are never re-typed.

SETUP (uses your existing Zoho app + the cron/GitHub infra you already run)
  pip install requests openpyxl
  Set env vars (Self Client or Server-based app, Leads READ scope: ZohoCRM.modules.leads.READ):
     ZOHO_DC            e.g. "com" (or "eu","in","com.au","ca","jp")
     ZOHO_CLIENT_ID
     ZOHO_CLIENT_SECRET
     ZOHO_REFRESH_TOKEN
  Then:
     python overnight_router.py --days 90 --out decision_log_90d.xlsx
     python overnight_router.py --date 2026-06-04 --out nightly.xlsx      # single night (cron)

NOTE ON CALIBRATION
  Source junk/positive rates are recomputed LIVE from the pulled window (self-tuning).
  The rep roster + trailing-7d load + shift start/end are a config snapshot below — refresh these
  from your Users/Shifts export (or a weekly job). Availability uses the round-robin flag; feed a
  real daily PTO list into REPS[...]['rr'] for exact "turn down vs catch up".
"""
import os, re, sys, time, argparse, datetime as dt
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PT = dt.timezone(dt.timedelta(hours=-7))  # PDT; engine treats the window in PT
OVERNIGHT_START_H = 18.0   # 6:00pm
OVERNIGHT_END_H   = 8.5    # 8:30am

# ---------------- CONFIG: rep roster (refresh from Users/Shifts export) ----------------
REPS_BASE=[
 {"name":"Russell Silva","office":"PT","rr":True,"load7":82},{"name":"Matthew Friedhoff","office":"PT","rr":True,"load7":47},
 {"name":"Eric Weiss","office":"PT","rr":True,"load7":66},{"name":"Samantha Riley","office":"PT","rr":True,"load7":67},
 {"name":"Brisa Green","office":"PT","rr":True,"load7":57},{"name":"Hovak Vartevanian","office":"PT","rr":False,"load7":0},
 {"name":"Cheryl Kopstein","office":"PT","rr":False,"load7":0},{"name":"Julian Landeros","office":"PT","rr":True,"load7":79},
 {"name":"Max Crozier","office":"ET","rr":True,"load7":79},{"name":"John Colovas","office":"ET","rr":True,"load7":83},
 {"name":"Jessie Todd","office":"ET","rr":True,"load7":59},{"name":"Frank Riffe","office":"ET","rr":True,"load7":72},
 {"name":"Melissa Daleske","office":"PT","rr":True,"load7":71},{"name":"Bryan Heflin","office":"PT","rr":True,"load7":76},
 {"name":"Faith Travis","office":"PT","rr":True,"load7":61},{"name":"Richard McCausland","office":"PT","rr":True,"load7":54},
 {"name":"Kirstan Villasenor","office":"PT","rr":True,"load7":55},{"name":"Julie Margolis","office":"PT","rr":True,"load7":71},
 {"name":"Daniel Jackson","office":"PT","rr":True,"load7":65},{"name":"Brandon Gilden","office":"PT","rr":True,"load7":83},
 {"name":"Jordan Koestner","office":"PT","rr":True,"load7":56},{"name":"Romario Carranza","office":"PT","rr":True,"load7":72},
 {"name":"Nathan Wilkie","office":"ET","rr":True,"load7":83},{"name":"Samuel Bender","office":"ET","rr":True,"load7":0},
]
EXCLUDE_NAMES={"Jacob Benner"}   # standing rule
REVIEW_ADMIN="House / SalesOps review queue"

POS={"Interested","Pre-Qualification Received","Sent Pre-Qualification to Merchant","Sent App to Merchant"}
SUS_TLD=(".site",".online",".store",".shop",".xyz",".top",".click",".group")
RISK_KW=["peptide","amino","sarm","research lab","chem","kratom","cbd","nicotine","vape","crypto","casino","gambling"]
TEST_CO={"test","dba name","business name","mayra01","tornado","hgjkh","sdsdf","needful things","ph"}
WEST_STATES={"CA","CALIFORNIA","WA","WASHINGTON","OR","OREGON","NV","NEVADA","AZ","ARIZONA","UT","UTAH","CO","COLORADO","ID","IDAHO","MT","MONTANA","WY","WYOMING","NM","NEW MEXICO","HI","HAWAII","AK","ALASKA"}
EAST_STATES={"NY","NEW YORK","FL","FLORIDA","NJ","NEW JERSEY","PA","PENNSYLVANIA","MA","MASSACHUSETTS","GA","GEORGIA","VA","VIRGINIA","NC","NORTH CAROLINA","SC","SOUTH CAROLINA","OH","OHIO","MI","MICHIGAN","IN","INDIANA","MD","MARYLAND","DC","DISTRICT OF COLUMBIA","DE","DELAWARE","CT","CONNECTICUT","ME","MAINE","NH","NEW HAMPSHIRE","VT","VERMONT","RI","RHODE ISLAND","WV","WEST VIRGINIA","TX","TEXAS","IL","ILLINOIS","MN","MINNESOTA","IA","IOWA","MO","MISSOURI","WI","WISCONSIN","KS","KANSAS","NE","NEBRASKA","OK","OKLAHOMA","AR","ARKANSAS","LA","LOUISIANA","MS","MISSISSIPPI","AL","ALABAMA","TN","TENNESSEE","ND","NORTH DAKOTA","SD","SOUTH DAKOTA","KY","KENTUCKY"}
WEST_AC=set("209 213 279 310 323 408 415 424 442 510 530 559 562 619 626 628 650 657 669 707 714 747 760 805 818 831 858 909 916 925 949 951 206 253 360 425 509 564 458 503 541 971 702 725 775 480 520 602 623 928 303 719 720 970 385 435 801 505 575 208 986 406 307 808 907".split())
EAST_AC=set("212 315 332 347 516 518 585 607 631 646 680 716 718 838 845 914 929 934 239 305 321 352 386 407 448 561 656 689 727 754 772 786 813 850 863 904 941 954 201 551 609 640 732 848 856 862 908 973 215 223 267 272 412 445 484 570 582 610 717 724 814 835 878 339 351 413 508 617 774 781 857 978 229 404 470 478 678 706 762 770 912 276 434 540 571 703 757 804 826 252 336 704 743 828 910 919 980 984 803 839 843 854 864 216 220 234 326 330 380 419 440 513 567 614 740 937 231 248 269 313 517 586 679 734 810 906 947 989 219 260 317 463 574 765 812 930 240 301 410 443 667 202 302 203 475 860 959 207 603 802 401 304 681 210 214 254 281 325 346 361 409 430 432 469 512 682 713 726 737 806 817 830 832 903 915 936 940 956 972 979 217 224 309 312 331 464 618 630 708 773 779 815 847 872 218 320 507 612 651 763 952 319 515 563 641 712 314 417 573 636 660 816 262 414 534 608 715 920 316 620 785 913 308 402 531 405 539 580 918 479 501 870 225 318 337 504 985 228 601 662 769 205 251 256 334 938 615 629 731 901 931 423 865 701 605 270 364 502 606 859".split())

def region_from(state,ac):
    if state:
        s=state.strip().upper()
        if s in WEST_STATES: return "West"
        if s in EAST_STATES: return "East"
    if ac:
        if ac in WEST_AC: return "West"
        if ac in EAST_AC: return "East"
    return "Unknown"
def phone_digits(*vals):
    for v in vals:
        if v:
            d=re.sub(r"\D","",str(v))
            if len(d)==11 and d.startswith("1"): d=d[1:]
            return d
    return ""
def phone_class(d):
    if d=="": return "empty"
    if len(d)==10 and d[0] in "23456789" and d[3] in "23456789" and len(set(d))>2: return "ok"
    if len(d)>10: return "foreign"
    return "invalid"
def gib(s):
    if not s: return False
    return bool(re.search(r"[bcdfghjklmnpqrstvwxz]{6,}",re.sub(r"[^a-z]","",s.lower())))

# ---------------- live source calibration ----------------
def build_source_stats(rows):
    tot={}; jk={}; ps={}
    for r in rows:
        s=r.get("source")
        if not s: continue
        tot[s]=tot.get(s,0)+1
        if r.get("actual")=="Junk Lead": jk[s]=jk.get(s,0)+1
        if r.get("actual") in POS: ps[s]=ps.get(s,0)+1
    return tot,jk,ps
class SrcStats:
    def __init__(s,tot,jk,ps): s.tot,s.jk,s.ps=tot,jk,ps
    def jr(s,src):
        t=s.tot.get(src); return (s.jk.get(src,0)/t) if t and t>=30 else None  # need volume to trust
    def pr(s,src):
        t=s.tot.get(src); return (s.ps.get(src,0)/t) if t and t>=30 else None
    def autohold(s): return {x for x in s.tot if (s.jr(x) or 0)>=0.75 and (s.pr(x) or 0)<=0.02}
    def priority(s): return {x for x in s.tot if (s.pr(x) or 0)>=0.10 and (s.jr(x) or 1)<=0.50}

# ---------------- routing ----------------
def fresh_reps():
    r=[dict(x) for x in REPS_BASE if x["name"] not in EXCLUDE_NAMES]
    for x in r: x["assigned"]=0
    return r
def pick_rep(reps,region):
    pool=[r for r in reps if r["rr"]]
    elig=[r for r in pool if r["office"]=="PT"] if region=="West" else pool
    if not elig: elig=pool
    elig.sort(key=lambda r:(r["assigned"],r["load7"],r["name"]))
    return elig[0]
def route_batch(rows,S):
    reps=fresh_reps(); seen={}; out=[]; AH=S.autohold(); PRI=S.priority()
    for ld in rows:
        email=(ld.get("email") or "").strip().lower(); company=ld.get("company") or ""
        name=f'{ld.get("first") or ""} {ld.get("last") or ""}'.strip(); src=ld.get("source")
        d=phone_digits(ld.get("phone"),ld.get("mobile")); pc=phone_class(d)
        ac=d[:3] if pc=="ok" else None; region=region_from(ld.get("state"),ac)
        blob=f"{company} {email}".lower(); risk=[k for k in RISK_KW if k in blob]
        is_test=name.lower() in ("test test","placeholder vm","placeholder","ph") or "miketestorg" in (src or "").lower() or "test@" in email or company.strip().lower() in TEST_CO
        sus=any(email.endswith(t) for t in SUS_TLD) or email.endswith(".co.uk") or gib(company)
        high=(S.jr(src) or 0)>=0.70; priority=src in PRI
        key=email or (d if pc=="ok" else ""); is_dup=bool(key) and key in seen
        flags=[]
        if pc=="foreign": flags.append("foreign/intl phone")
        elif pc=="invalid": flags.append("invalid phone")
        elif pc=="empty": flags.append("no phone")
        if risk: flags.append("restricted vertical")
        if sus: flags.append("suspicious email/co.")
        if is_dup: flags.append("dup in batch")
        if key and key not in seen: seen[key]=name
        if is_dup: dec,who,why="Duplicate \u2192 merge to original","(original owner)","Same email/phone already in batch."
        elif is_test: dec,who,why="Obvious junk \u2192 Admin review",REVIEW_ADMIN,"Test/placeholder/dummy record."
        elif risk: dec,who,why="Restricted vertical \u2192 Admin review",REVIEW_ADMIN,f"High-risk vertical ({', '.join(risk)})."
        elif pc=="foreign": dec,who,why="International \u2192 Admin review",REVIEW_ADMIN,"Non-US phone; outside US merchant base."
        elif src in AH: dec,who,why="Auto-hold source \u2192 Admin review",REVIEW_ADMIN,f"Source '{src}' {S.jr(src):.0%} junk / {S.pr(src):.0%} positive."
        elif pc in ("empty","invalid"):
            if email: dec,who,why="No valid phone \u2192 auto-status 'Bad Number/Good Email'","(email nurture)","No dialable US number; email present."
            else: dec,who,why="No contact \u2192 auto-status 'Needs Enrichment'","(enrichment)","No phone or email."
        elif sus and high: dec,who,why="Likely junk \u2192 Admin review",REVIEW_ADMIN,f"Suspicious + high-junk source ({src})."
        else:
            rep=pick_rep(reps,region); rep["assigned"]+=1; who=rep["name"]
            dec="Assign to rep"+(" (PRIORITY)" if priority else "")
            tz={"West":"West \u2192 Encino","East":"East/Central \u2192 any","Unknown":"Unknown \u2192 flexible"}[region]
            why=f"Clean US contact. {tz}. Even-split parity."+(" PRIORITY source." if priority else "")
        out.append({**ld,"region":region,"decision":dec,"assignee":who,"reason":why,"flags":"; ".join(flags),"priority":priority})
    return out

# ---------------- Zoho ----------------
def access_token():
    dc=os.environ.get("ZOHO_DC","com")
    r=requests.post(f"https://accounts.zoho.{dc}/oauth/v2/token",data={
        "refresh_token":os.environ["ZOHO_REFRESH_TOKEN"],"client_id":os.environ["ZOHO_CLIENT_ID"],
        "client_secret":os.environ["ZOHO_CLIENT_SECRET"],"grant_type":"refresh_token"})
    r.raise_for_status(); return r.json()["access_token"]
def coql(token, query):
    dc=os.environ.get("ZOHO_DC","com")
    rows=[]; offset=0
    while True:
        q=query+f" LIMIT {offset},200"
        r=requests.post(f"https://www.zohoapis.{dc}/crm/v6/coql",
                        headers={"Authorization":f"Zoho-oauthtoken {token}"},json={"select_query":q})
        if r.status_code==204: break
        r.raise_for_status(); j=r.json(); rows+=j.get("data",[])
        if not j.get("info",{}).get("more_records"): break
        offset+=200; time.sleep(0.2)
    return rows
def pull_leads(token,start,end):
    q=(f"SELECT id, First_Name, Last_Name, Company, Email, Phone, Mobile, State, Lead_Source, "
       f"Lead_Status, Created_Time FROM Leads WHERE Created_Time between '{start}' and '{end}'")
    raw=coql(token,q)
    out=[]
    for r in raw:
        out.append({"first":r.get("First_Name"),"last":r.get("Last_Name"),"company":r.get("Company"),
                    "email":r.get("Email"),"phone":r.get("Phone"),"mobile":r.get("Mobile"),
                    "state":r.get("State"),"source":r.get("Lead_Source"),"actual":r.get("Lead_Status"),
                    "created":r.get("Created_Time")})
    return out

def is_overnight(created_iso):
    try: t=dt.datetime.fromisoformat(created_iso).astimezone(PT)
    except Exception: return False,None
    h=t.hour+t.minute/60.0
    overnight = (h>=OVERNIGHT_START_H) or (h<OVERNIGHT_END_H)
    night_date = (t.date()+dt.timedelta(days=1)) if h>=OVERNIGHT_START_H else t.date()  # label by the morning
    return overnight, night_date.isoformat()

# ---------------- Excel ----------------
def write_xlsx(decisions, summary, path):
    FNT="Arial"; HF=PatternFill("solid",fgColor="1F3864"); HFONT=Font(name=FNT,bold=True,color="FFFFFF",size=10)
    thin=Side(style="thin",color="BFBFBF"); BD=Border(thin,thin,thin,thin)
    WRAP=Alignment(wrap_text=True,vertical="top"); CEN=Alignment(horizontal="center",vertical="center")
    F_A=PatternFill("solid",fgColor="E2EFDA");F_R=PatternFill("solid",fgColor="FCE4D6");F_AU=PatternFill("solid",fgColor="FFF2CC");F_D=PatternFill("solid",fgColor="EDEDED");F_P=PatternFill("solid",fgColor="DDEBF7")
    def H(ws,n):
        for c in range(1,n+1):
            x=ws.cell(row=1,column=c); x.fill=HF; x.font=HFONT; x.alignment=CEN; x.border=BD
    wb=Workbook(); ws=wb.active; ws.title="Lead Decisions"; ws.sheet_view.showGridLines=False
    hd=["Date","Name","Company","Phone","State","Region","Source","Decision","Assigned to","Why","Flags","Eventual status"]
    ws.append(hd); H(ws,len(hd)); ws.freeze_panes="A2"
    for i,l in enumerate(decisions,1):
        ws.append([l["date"],f'{l.get("first") or ""} {l.get("last") or ""}'.strip(),l.get("company") or "",
                   (l.get("phone") or l.get("mobile") or ""),l.get("state") or "",l["region"],l.get("source") or "",
                   l["decision"],l["assignee"],l["reason"],l["flags"],l.get("actual") or ""])
        dec=l["decision"]; fill=F_A if dec.startswith("Assign") else F_R if "review" in dec else F_AU if "auto-status" in dec else F_D if dec.startswith("Duplicate") else F_R
        if l["priority"]: fill=F_P
        for c in range(1,len(hd)+1):
            x=ws.cell(row=i+1,column=c); x.border=BD; x.font=Font(name=FNT,size=9); x.alignment=WRAP
            if c==8: x.fill=fill
    for col,w in zip("ABCDEFGHIJKL",[11,17,20,13,7,8,15,28,18,38,20,18]): ws.column_dimensions[col].width=w
    ws2=wb.create_sheet("Night Summary"); ws2.sheet_view.showGridLines=False
    hd2=["Night","Leads","Assigned","Review","Auto-status","Dups","Priority","Eventual junk","Junk caught","Junk leaked","Good held"]
    ws2.append(hd2); H(ws2,len(hd2)); ws2.freeze_panes="A2"
    for i,row in enumerate(summary,1):
        ws2.append(row)
        for c in range(1,len(hd2)+1):
            x=ws2.cell(row=i+1,column=c); x.border=BD; x.font=Font(name=FNT,size=9); x.alignment=CEN
    for col,w in zip("ABCDEFGHIJK",[12,8,10,9,12,7,9,13,12,12,11]): ws2.column_dimensions[col].width=w
    wb.save(path)

def summarize(night, dec):
    a=sum(1 for d in dec if d["decision"].startswith("Assign"))
    rev=sum(1 for d in dec if "Admin review" in d["decision"]); au=sum(1 for d in dec if "auto-status" in d["decision"])
    du=sum(1 for d in dec if d["decision"].startswith("Duplicate")); pri=sum(1 for d in dec if d["priority"] and d["decision"].startswith("Assign"))
    tj=sum(1 for d in dec if d.get("actual")=="Junk Lead")
    held=[d for d in dec if not d["decision"].startswith("Assign")]
    caught=sum(1 for d in held if d.get("actual")=="Junk Lead")
    leaked=sum(1 for d in dec if d["decision"].startswith("Assign") and d.get("actual")=="Junk Lead")
    fp=sum(1 for d in held if d.get("actual") in POS)
    return [night,len(dec),a,rev,au,du,pri,tj,caught,leaked,fp]

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--days",type=int); ap.add_argument("--date"); ap.add_argument("--out",default="decision_log.xlsx")
    a=ap.parse_args()
    if a.date:
        end=dt.datetime.fromisoformat(a.date+"T08:30:00").replace(tzinfo=PT)
        start=end-dt.timedelta(hours=14.5)
    else:
        days=a.days or 90; end=dt.datetime.now(PT); start=end-dt.timedelta(days=days)
    token=access_token()
    leads=pull_leads(token, start.isoformat(timespec="seconds"), end.isoformat(timespec="seconds"))
    # keep overnight only, tag night
    on=[]
    for l in leads:
        ok,nd=is_overnight(l.get("created") or "")
        if ok: l["date"]=nd; on.append(l)
    S=SrcStats(*build_source_stats(leads))   # calibrate on the FULL window for stable rates
    by_night={}
    for l in on: by_night.setdefault(l["date"],[]).append(l)
    all_dec=[]; summ=[]
    for night in sorted(by_night):
        dec=route_batch(by_night[night],S); all_dec+=dec; summ.append(summarize(night,dec))
    write_xlsx(all_dec,summ,a.out)
    print(f"{len(leads)} leads pulled, {len(on)} overnight, {len(by_night)} nights -> {a.out}")

if __name__=="__main__":
    main()
